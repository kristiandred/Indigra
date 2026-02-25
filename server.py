#!/usr/bin/env python3
import json
import os
import subprocess
from http.server import HTTPServer, SimpleHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
import openpyxl
from openpyxl.utils import get_column_letter

class AdminAPIHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        self.data_dir = os.path.join(os.getcwd(), 'data')
        super().__init__(*args, **kwargs)
    
    def do_GET(self):
        parsed_path = urlparse(self.path)
        
        if parsed_path.path.startswith('/api/'):
            self.handle_api_get(parsed_path)
        else:
            # Обслуживаем статические файлы
            super().do_GET()
    
    def do_POST(self):
        parsed_path = urlparse(self.path)
        
        if parsed_path.path.startswith('/api/'):
            self.handle_api_post(parsed_path)
        else:
            self.send_error(404)
    
    def do_DELETE(self):
        parsed_path = urlparse(self.path)
        
        if parsed_path.path.startswith('/api/'):
            self.handle_api_delete(parsed_path)
        else:
            self.send_error(404)
    
    def handle_api_get(self, parsed_path):
        if parsed_path.path == '/api/files':
            self.get_files()
        elif parsed_path.path.startswith('/api/file/'):
            filename = parsed_path.path.split('/')[-1]
            self.get_file(filename)
        else:
            self.send_error(404)
    
    def handle_api_post(self, parsed_path):
        if parsed_path.path.startswith('/api/file/'):
            filename = parsed_path.path.split('/')[-1]
            self.save_file(filename)
        elif parsed_path.path == '/api/file':
            self.create_file()
        else:
            self.send_error(404)
    
    def handle_api_delete(self, parsed_path):
        if parsed_path.path.startswith('/api/file/'):
            filename = parsed_path.path.split('/')[-1]
            self.delete_file(filename)
        else:
            self.send_error(404)
    
    def get_files(self):
        try:
            if not os.path.exists(self.data_dir):
                os.makedirs(self.data_dir)
            
            files = [f for f in os.listdir(self.data_dir) if f.endswith('.xlsx')]
            self.send_json_response(files)
        except Exception as e:
            self.send_error(500, str(e))
    
    def get_file(self, filename):
        try:
            file_path = os.path.join(self.data_dir, filename)
            if not os.path.exists(file_path):
                self.send_error(404, f"Файл {filename} не найден")
                return
            
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            data = []
            
            for row in sheet.iter_rows(values_only=True):
                data.append(list(cell if cell is not None else '' for cell in row))
            
            self.send_json_response(data)
        except Exception as e:
            self.send_error(500, str(e))
    
    def save_file(self, filename):
        try:
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            data = json.loads(post_data.decode('utf-8'))
            
            file_path = os.path.join(self.data_dir, filename)
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            
            for row_idx, row in enumerate(data['data'], 1):
                for col_idx, value in enumerate(row, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)
            
            workbook.save(file_path)
            self.git_commit(f"Обновлен файл {filename}")
            self.send_json_response({"success": True, "message": f"Файл {filename} сохранен"})
        except Exception as e:
            self.send_error(500, str(e))
    
    def create_file(self):
        try:
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            data = json.loads(post_data.decode('utf-8'))
            
            filename = data['filename']
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            file_path = os.path.join(self.data_dir, filename)
            if os.path.exists(file_path):
                self.send_error(400, "Файл уже существует")
                return
            
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            
            for row_idx, row in enumerate(data['data'], 1):
                for col_idx, value in enumerate(row, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)
            
            workbook.save(file_path)
            self.git_commit(f"Создан файл {filename}")
            self.send_json_response({"success": True, "message": f"Файл {filename} создан"})
        except Exception as e:
            self.send_error(500, str(e))
    
    def delete_file(self, filename):
        try:
            file_path = os.path.join(self.data_dir, filename)
            if not os.path.exists(file_path):
                self.send_error(404, f"Файл {filename} не найден")
                return
            
            os.remove(file_path)
            self.git_commit(f"Удален файл {filename}")
            self.send_json_response({"success": True, "message": f"Файл {filename} удален"})
        except Exception as e:
            self.send_error(500, str(e))
    
    def git_commit(self, message):
        try:
            subprocess.run(['git', 'add', 'data/'], cwd=os.getcwd(), check=True, capture_output=True)
            subprocess.run(['git', 'commit', '-m', message], cwd=os.getcwd(), check=True, capture_output=True)
            try:
                subprocess.run(['git', 'push'], cwd=os.getcwd(), check=True, capture_output=True)
            except:
                pass  # Игнорируем ошибки push
        except:
            pass  # Игнорируем ошибки git
    
    def send_json_response(self, data):
        self.send_response(200)
        self.send_header('Content-type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode('utf-8'))
    
    def end_headers(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, DELETE, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        super().end_headers()

def run_server():
    port = 3000
    server = HTTPServer(('localhost', port), AdminAPIHandler)
    print(f"Сервер запущен на порту {port}")
    print(f"Откройте http://localhost:{port}/admin-panel.html")
    server.serve_forever()

if __name__ == '__main__':
    run_server()
